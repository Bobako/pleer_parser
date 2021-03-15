import openpyxl
import json
import time



def to_json(dict_):						#converts dict to json string
	return json.dumps(dict_)

def to_dict(json_string):
	return json.loads(json_string)		#converts json string to dict

def read_json(file_name):          		#converts json string from file to dict
	file = open(file_name,'r',encoding = "utf-8")
	json_string = file.read()
	file.close()
	try:
		dict_ = json.loads(json_string)
	except Exception:
		dict_ = {}
	return dict_	

def write_json(file_name,dict_):        #writes dict to json file (if file doesn't exist creates new one, else adds dict's keys and literals)
										#if some key already exist, it's literal will be replaced with a new one
										#KEYS MUSTN'T BE INT
	try:
		file = open(file_name,'x')
		file.close()

	except Exception:					#if file exist

		file = open(file_name,'r')
		json_string_exist = file.read()
		file.close()

		dict_exist = json.loads(json_string_exist)

		dict_  = {**dict_exist,**dict_}

	finally:
		json_string = json.dumps(dict_)
		file = open(file_name,'w')
		file.write(json_string)
		file.close()


def remove_json_key(file_name,key):
	with open (file_name,'r') as f:
		dict_ = json.loads(f.read())
	try:
		a = dict_.pop(key)
	except Exception:
		pass
	with open (file_name,'w') as f:
		f.write(json.dumps(dict_))


		
def read_xlsx(file_name):				#converts xlsx to dict, names of sheets and columns (the first row cells) became dict keys
	
	file = openpyxl.load_workbook(file_name)

	keys = []
	book_dict = {}
	sheet_dict = {}
	row_dict = {}

	for sheetname in file.sheetnames:
		sheet = file[sheetname]
		keys = columns_names(sheet)
		sheet_dict = {}

		for i in range(2,rows_number(sheet)+1):
			row_dict = {}
			for j in range(1,len(keys)+1):
				row_dict[keys[j-1]] = sheet.cell(row = i, column = j).value
			sheet_dict[str(i-1)] = row_dict

		book_dict[sheetname] = sheet_dict

	return book_dict


		
def add_row_xlsx(file_name,sheetname,dict_):		#converts dict to a table row and adds it to xlsx file
	file = openpyxl.load_workbook(file_name)
	sheet = file[sheetname]
	keys = columns_names(sheet)
	literals = []
	l = rows_number(sheet)+1
	for key in keys:
		literals.append(dict_[key])

	for i in range(1,len(keys)+1):
		sheet.cell(row = l, column = i, value = str(literals[i-1]))

	file.save(file_name)

def add_sheet(file_name,sheetname):					#if there is not sheet with name sheetname in a xlsx book, adds it and returns True, else returns False
	file = openpyxl.load_workbook(file_name)
	if sheetname in file.sheetnames:
		return False
	else:
		file.create_sheet(sheetname)
		file.save(file_name)
		write_sheet_prefab(file_name,sheetname)
		return True


def write_sheet_prefab(file_name,sheetname):
	file = openpyxl.load_workbook(file_name)
	sheet = file[sheetname]
	keys = columns_names(openpyxl.load_workbook(file_name)["example"])

	for i in range(len(keys)):
		sheet.cell(row = 1, column = i+1, value = str(keys[i]))

	file.save(file_name)


def write_log(log_file,log):            			#writes a line (in format: date-time-log string) to log_file
	file = open(log_file, 'a')
	
	tm = time.localtime()
	tm = time.strftime('%d.%m.%Y %X',tm)

	file.write(tm+ ' '+ log+'\n')

	file.close()


def read_random_ids(file_name):                      #returns all random_ids from file (for vk chat bot)
	file = open(file_name, 'r')
	random_ids_str = file.read()
	file.close()

	random_ids = random_ids_str.split(';')[:-1:]
	for i in range (len(random_ids)):
		try:
			random_ids[i] = int(random_ids[i])
		except Exception:
			random_id = 0

	return random_ids

def write_random_id(random_id,file_name):			#adds a new random_id to file and removes an older one (for vk chat bot)
	random_ids = read_random_ids(file_name)
	if len(random_ids)>100:
		random_ids  = random_ids[1::]
	
	random_ids = random_ids+[random_id]
	random_ids_str = ''
	
	for random_id_ in random_ids:
		random_ids_str += (str(random_id_) + ';')

	file = open(file_name,'w')
	file.write(random_ids_str)
	file.close()


def get_parts(file_name = 'db_v2.xlsx',sheet_name = "Sheet1"):				#gets parts from bd (by mal)(for amtech vk bot )
	book = read_xlsx(file_name)
	stook = book[sheet_name]
	parts = []
	MAXIM = len(stook)
	for row in range (1,MAXIM+1):
		row = str(row)
		line = stook[row]
		parts.append(line)
	return(parts)













#funcs for read_xlsx
def columns_names(sheet):
	i = 1
	keys = []
	while sheet.cell(row = 1, column = i).value!=None:
		keys.append(str(sheet.cell(row = 1, column = i).value))
		i+=1
	return keys
def rows_number(sheet):
	i=1
	while sheet.cell(row = i, column = 1).value!=None:
		i+=1
	return i-1




	








	


#print (read_json('settings.json')["test"])


#print(read_xlsx("test.xlsx"))


#file = openpyxl.load_workbook("test.xlsx")
#i = int(input())
#test_dict = {"Личность":"Она",
#			 "Амудешник?":"Нет",
#			 "Пидор?":"Девка"}

#add_row_xlsx("test.xlsx","Лист1",test_dict )


#write_log("log.txt","Хочу сказать этому обамэ ебамана")


#dict_ = read_xlsx('test.xlsx')
#print(dict_)




